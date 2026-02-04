import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from stock_list_fetcher import get_stock_list_with_fallback

def create_smart_template():
    """종목명 선택 시 종목코드가 자동으로 입력되는 엑셀 템플릿 생성"""

    # 전체 상장 종목 리스트 가져오기 (4000개 이상)
    print("전체 상장 종목 로딩 중...")
    stock_data = get_stock_list_with_fallback()
    stock_list = [(s['name'], s['code']) for s in stock_data]
    print(f"{len(stock_list)}개 종목 로드 완료")


    # 워크북 생성
    wb = Workbook()

    # Sheet 1: 입력 시트
    ws_input = wb.active
    ws_input.title = "포트폴리오 입력"

    # 헤더 설정
    headers = ['종목명', '종목코드', '평균단가', '보유수량', '매입금액']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws_input.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 컬럼 너비 조정
    ws_input.column_dimensions['A'].width = 20
    ws_input.column_dimensions['B'].width = 12
    ws_input.column_dimensions['C'].width = 15
    ws_input.column_dimensions['D'].width = 15
    ws_input.column_dimensions['E'].width = 18

    # Sheet 2: 종목 코드표
    ws_codes = wb.create_sheet("종목코드표")
    ws_codes['A1'] = '종목명'
    ws_codes['B1'] = '종목코드'
    ws_codes['A1'].font = header_font
    ws_codes['B1'].font = header_font
    ws_codes['A1'].fill = header_fill
    ws_codes['B1'].fill = header_fill

    # 종목 리스트 입력
    for idx, (name, code) in enumerate(stock_list, start=2):
        ws_codes[f'A{idx}'] = name
        ws_codes[f'B{idx}'] = code

    ws_codes.column_dimensions['A'].width = 20
    ws_codes.column_dimensions['B'].width = 12

    # 데이터 유효성 검사 (종목명 드롭다운 - 직접 입력도 허용)
    dv = DataValidation(
        type="list",
        formula1=f'종목코드표!$A$2:$A${len(stock_list)+1}',
        allow_blank=True,
        showErrorMessage=False  # 오류 메시지 표시 안 함 (자유 입력 허용)
    )
    dv.prompt = '드롭다운에서 선택하거나 직접 입력하세요'
    dv.promptTitle = '종목명 입력'
    dv.error = '목록에 없는 종목도 입력 가능합니다'
    dv.errorTitle = '안내'
    ws_input.add_data_validation(dv)

    # A2:A100까지 데이터 유효성 검사 적용
    for row in range(2, 102):
        dv.add(f'A{row}')

    # VLOOKUP 수식 추가 (종목코드 자동 입력)
    for row in range(2, 102):
        # B열(종목코드)에 VLOOKUP 수식
        ws_input[f'B{row}'] = f'=IFERROR(VLOOKUP(A{row},종목코드표!$A$2:$B${len(stock_list)+1},2,FALSE),"")'
        # E열(매입금액)에 자동 계산 수식 (평균단가 × 보유수량)
        ws_input[f'E{row}'] = f'=IF(AND(C{row}<>"",D{row}<>""),C{row}*D{row},"")'

    # Sheet 3: 사용 설명서
    ws_guide = wb.create_sheet("📖 사용방법", 0)  # 첫 번째 시트로
    ws_guide['A1'] = '📖 스마트 포트폴리오 템플릿 사용 방법'
    ws_guide['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_guide['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_guide.merge_cells('A1:D1')

    guide_text = [
        ('', ''),
        ('✨ 종목명 입력 방법 (2가지)', ''),
        ('', ''),
        ('방법 1️⃣ 드롭다운 선택', '▶ A열 클릭 → 드롭다운 화살표 클릭 → 종목 선택'),
        ('방법 2️⃣ 직접 입력', '▶ A열에 종목명 직접 타이핑 (예: 삼성전자)'),
        ('', '▶ 목록에 없는 종목도 입력 가능!'),
        ('', ''),
        ('💡 종목코드 자동 입력', ''),
        ('', '▶ 종목명을 입력하면 B열(종목코드)이 자동으로 입력됩니다'),
        ('', '▶ 목록에 없는 종목은 종목코드를 직접 입력하세요'),
        ('', ''),
        ('📝 입력 항목', ''),
        ('', '• 종목명: 주식 이름 (드롭다운 또는 직접 입력)'),
        ('', '• 종목코드: 자동 입력 (수동 입력도 가능)'),
        ('', '• 평균단가: 매입한 평균 가격 (원)'),
        ('', '• 보유수량: 보유한 주식 수 (주)'),
        ('', '• 매입금액: 자동 계산 (수동 수정 가능)'),
        ('', '  → 평균단가 × 보유수량으로 자동 계산'),
        ('', '  → 실제 매입금액이 다르면 직접 수정하세요!'),
        ('', ''),
        ('🎯 작성 후', ''),
        ('', '1. 파일 저장'),
        ('', '2. 대시보드에서 업로드'),
        ('', '3. 일괄 등록 버튼 클릭!'),
        ('', ''),
        ('📌 주요 종목 100개 이상 포함 (종목코드표 시트 참고)', ''),
    ]

    for idx, (col1, col2) in enumerate(guide_text, start=2):
        ws_guide[f'A{idx}'] = col1
        ws_guide[f'B{idx}'] = col2
        if col1.startswith('방법') or col1.startswith('💡') or col1.startswith('📝') or col1.startswith('🎯'):
            ws_guide[f'A{idx}'].font = Font(bold=True, size=11)

    ws_guide.column_dimensions['A'].width = 30
    ws_guide.column_dimensions['B'].width = 50

    # 샘플 데이터 추가
    ws_input['A2'] = '삼성전자'
    ws_input['C2'] = 70000
    ws_input['D2'] = 10

    ws_input['A3'] = 'SK하이닉스'
    ws_input['C3'] = 120000
    ws_input['D3'] = 5

    # 안내 메시지 추가
    ws_input['A105'] = '💡 종목명을 드롭다운에서 선택하거나 직접 입력하세요. 종목코드는 자동으로 입력됩니다!'
    ws_input['A105'].font = Font(italic=True, color="0000FF")
    ws_input.merge_cells('A105:D105')

    # 바이트 스트림으로 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


def get_stock_code_dict():
    """종목명-종목코드 딕셔너리 반환"""
    stock_list = [
        ('삼성전자', '005930'),
        ('SK하이닉스', '000660'),
        ('현대차', '005380'),
        ('기아', '000270'),
        ('POSCO홀딩스', '005490'),
        ('NAVER', '035420'),
        ('카카오', '035720'),
        ('삼성바이오로직스', '207940'),
        ('LG화학', '051910'),
        ('삼성SDI', '006400'),
        # ... 더 많은 종목들
    ]
    return dict(stock_list)
